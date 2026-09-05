import json
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime, date, timezone as dt_timezone

from django.contrib.sites.models import Site
from django.core.management.base import BaseCommand, CommandError
from django.db.models import Count, GeneratedField
from django.db.models.functions import Lower
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from gregory.models import OrganizationSite, Trials, Subject, TeamCategory
from sitesettings.models import CustomSetting


EXCLUDED_SCALARS = frozenset({"utitle", "usummary"})
EXCLUDED_M2M = frozenset({"ml_predictions"})

# Ordered column groups for data sheets
IDENTITY_COLS = [
	"trial_id",
	"title",
	"acronym",
	"scientific_title",
	"link",
	"discovery_date",
	"last_updated",
	"published_date",
	"date_registration",
	"last_refreshed_on",
	"export_date",
]

SCALAR_ORDER = [
	"summary",
	"internal_number",
	"secondary_id",
	"source_register",
	"other_records",
	"prospective_registration",
	"study_type",
	"study_type_normalized",
	"study_design",
	"phase",
	"phase_normalized",
	"recruitment_status",
	"recruitment_status_normalized",
	"target_size",
	"date_enrollement",
	"countries",
	"countries_by_source",
	"regions_normalized",
	"condition",
	"intervention",
	"primary_outcome",
	"secondary_outcome",
	"inclusion_criteria",
	"exclusion_criteria",
	"inclusion_agemin",
	"inclusion_agemax",
	"inclusion_gender",
	"inclusion_gender_normalized",
	"primary_sponsor",
	"secondary_sponsor",
	"source_support",
	"sponsor_type",
	"lead_sponsor_class",
	"contact_firstname",
	"contact_lastname",
	"contact_address",
	"contact_email",
	"contact_tel",
	"contact_affiliation",
	"ethics_review_status",
	"ethics_review_approval_date",
	"ethics_review_contact_name",
	"ethics_review_contact_address",
	"ethics_review_contact_phone",
	"ethics_review_contact_email",
	"results_posted",
	"results_date_completed",
	"results_url_link",
	"results_yes_no",
	"results_ipd_plan",
	"results_ipd_description",
	"therapeutic_areas",
	"country_status",
	"trial_region",
	"overall_decision_date",
	"countries_decision_date",
	"ctg_detailed_description",
]

RELATION_COLS = [
	"subjects",
	"teams",
	"sources",
	"team_categories",
	"articles",
	"trial_countries",
	"sponsor_id",
	"sponsor_slug",
	"primary_sponsor_normalized",
	"sponsor_type_normalized",
	"sponsor_type_source",
]

# Descriptions for exported columns absent from TrialAdminForm.Meta.help_texts.
# Format: field_name → (label, description, source_registries)
EXTRA_GLOSSARY = {
	"trial_id": (
		"Trial ID",
		"Internal Gregory database identifier for this trial record.",
		"",
	),
	"summary": (
		"Summary",
		"Plain-language summary of the trial.",
		"WHO ICTRP, ClinicalTrials.gov, EU CTIS",
	),
	"discovery_date": (
		"Discovery date",
		"Date this trial was first added to Gregory.",
		"",
	),
	"last_updated": (
		"Last updated",
		"Date and time this record was last modified in Gregory.",
		"",
	),
	"identifiers_json": (
		"Identifiers (raw JSON)",
		"Raw JSON dict of all registry identifiers for this trial.",
		"WHO ICTRP, ClinicalTrials.gov, EU CTIS",
	),
	"subjects": (
		"Subjects",
		"Research subjects this trial is assigned to in Gregory (semicolon-separated).",
		"",
	),
	"teams": (
		"Teams",
		"Teams this trial is assigned to in Gregory (semicolon-separated).",
		"",
	),
	"sources": (
		"Sources",
		"Registry sources that provided data for this trial (semicolon-separated).",
		"",
	),
	"team_categories": (
		"Categories",
		"Team categories assigned to this trial (semicolon-separated). See the "
		"Categories sheet for each category's description and search terms.",
		"",
	),
	"articles": (
		"Related articles",
		"Count of articles that reference this trial, followed by their URLs.",
		"",
	),
	"therapeutic_areas": (
		"Therapeutic areas",
		"Therapeutic areas covered by the trial.",
		"EU CTIS",
	),
	"country_status": (
		"Country status",
		"Authorisation status of the trial in each participating country.",
		"EU CTIS",
	),
	"countries_by_source": (
		"Countries by source",
		"Raw per-source country lists as a JSON map keyed by registry slug "
		"(ctgov, ictrp). Each importer writes only its own key, so sources no longer "
		"overwrite each other's country data (the flat countries column is last-writer-wins).",
		"WHO ICTRP, ClinicalTrials.gov",
	),
	"regions_normalized": (
		"Regions (normalized)",
		"Continental regions derived from the trial's normalized countries "
		"(africa, asia, europe, north_america, south_america, oceania), plus any "
		"literal region tokens found in the raw country data.",
		"WHO ICTRP, ClinicalTrials.gov, EU CTIS",
	),
	"trial_countries": (
		"Countries (normalized)",
		"Per-country breakdown: display name and ISO 3166-1 alpha-2 code, with EU CTIS "
		"authorisation status, decision date, and contributing source slugs where known. "
		'Format: "Germany [DE] (recruiting; 2024-07-19; src: ctgov+ctis); …".',
		"WHO ICTRP, ClinicalTrials.gov, EU CTIS",
	),
	"trial_region": ("Trial region", "Geographic region of the trial.", "EU CTIS"),
	"overall_decision_date": (
		"Overall decision date",
		"Date of the overall regulatory decision for the trial.",
		"EU CTIS",
	),
	"countries_decision_date": (
		"Countries decision dates",
		"JSON map of per-country regulatory decision dates.",
		"EU CTIS",
	),
	"sponsor_type": (
		"Sponsor type (raw)",
		"Sponsor category as reported verbatim by EU CTIS (e.g. \"Pharmaceutical "
		"company\"). Populated for EU CTIS trials only — see sponsor_type_normalized "
		"for a canonical category derived across all three registries.",
		"EU CTIS",
	),
	"lead_sponsor_class": (
		"Lead sponsor class (CTGov)",
		"Sponsor agency class as classified by ClinicalTrials.gov (e.g. INDUSTRY, NIH, "
		"FED, OTHER_GOV, INDIV, NETWORK, OTHER, AMBIG, UNKNOWN). One of the signals "
		"feeding sponsor_type_normalized.",
		"ClinicalTrials.gov",
	),
	"sponsor_id": (
		"Sponsor ID (canonical)",
		"Database id of the canonical sponsor entity — stable across spelling-variant "
		"merges, so it's the reliable join key for grouping/joining against the "
		"/sponsors/ API endpoint. Blank when primary_sponsor is empty.",
		"WHO ICTRP, ClinicalTrials.gov, EU CTIS",
	),
	"sponsor_slug": (
		"Sponsor slug (canonical)",
		"URL-safe slug of the canonical sponsor entity, as used by the /sponsors/ API "
		"endpoint's sponsor_slug filter. Blank when primary_sponsor is empty.",
		"WHO ICTRP, ClinicalTrials.gov, EU CTIS",
	),
	"primary_sponsor_normalized": (
		"Sponsor (canonical)",
		"Canonical sponsor entity name. Spelling variants of the same real-world "
		"sponsor across registries (e.g. \"Novartis\", \"Novartis Pharma AG\", "
		"\"Novartis Pharmaceuticals\") resolve to one name here, so counting/grouping "
		"trials by sponsor no longer undercounts due to spelling differences. Blank "
		"when primary_sponsor is empty.",
		"WHO ICTRP, ClinicalTrials.gov, EU CTIS",
	),
	"sponsor_type_normalized": (
		"Sponsor type (canonical)",
		"Canonical sponsor category for the resolved sponsor entity: industry, "
		"academic_medical, government, nonprofit, or other. Derived from (in priority "
		"order) a curated hand-assignment for known sponsor families, "
		"ClinicalTrials.gov's lead_sponsor_class, EU CTIS's raw sponsor_type, or "
		"keyword rules on the sponsor name — so it is populated far more often than "
		"the raw sponsor_type column, which only EU CTIS provides. See "
		"sponsor_type_source for which of these actually applied. Blank when no "
		"signal was available to classify the sponsor.",
		"WHO ICTRP, ClinicalTrials.gov, EU CTIS",
	),
	"sponsor_type_source": (
		"Sponsor type source",
		"Audit trail for sponsor_type_normalized: which signal actually determined it — "
		"\"curated\" (set by hand for a known sponsor family, never overwritten "
		"automatically), \"ctgov\" (ClinicalTrials.gov's lead_sponsor_class), \"ctis\" "
		"(EU CTIS's raw sponsor_type), or \"rules\" (keyword match on the sponsor name, "
		"the lowest-confidence tier). Blank when sponsor_type_normalized itself is blank.",
		"WHO ICTRP, ClinicalTrials.gov, EU CTIS",
	),
	"ctg_detailed_description": (
		"Detailed description",
		"Extended description from ClinicalTrials.gov.",
		"ClinicalTrials.gov",
	),
}

REGISTRIES_OVERVIEW = [
	(
		"WHO ICTRP",
		"Varies (nct, euctr, chictr, nl, …)",
		"Aggregator of national/regional registries worldwide",
		"Richest field coverage: ethics review, IPD plans, enrolment dates, full sponsor "
		"info. May lag primary registries; can overwrite a field with an empty value on "
		"re-import.",
	),
	(
		"ClinicalTrials.gov",
		"nct",
		"US-hosted global registry",
		"Provides ctg_detailed_description and results_url_link. Combines exclusion into "
		"inclusion criteria. Never blanks a field on update (non-destructive).",
	),
	(
		"EU Clinical Trials (CTIS)",
		"euctr / eudract / ctis",
		"EU trials register",
		"Provides therapeutic_areas, country_status, trial_region, overall_decision_date, "
		"countries_decision_date, sponsor_type. Does not provide date_registration. "
		"May overwrite fields with empty values on re-import.",
	),
]

MERGE_PROSE = (
	"A single trial can be ingested from more than one source registry. Gregory stores "
	"one row per trial and merges data on re-import using a last-write-wins strategy. "
	"ClinicalTrials.gov never blanks a field it previously populated; WHO ICTRP and EU "
	"CTIS may overwrite existing values — including with empty ones — if the incoming "
	"record omits a field. The identifiers column is always merged non-destructively: "
	"keys are added, never removed. Fields produced by only one registry (e.g. EU CTIS "
	"therapeutic_areas or CT.gov ctg_detailed_description) are set only by their "
	"respective importer and are never in conflict."
)

REGISTRY_NAMES = ["WHO ICTRP", "ClinicalTrials.gov", "EU CTIS"]

CATEGORY_COLUMNS = [
	"Subject",
	"Team",
	"Category",
	"Slug",
	"Description",
	"Search terms",
	"Terms (count)",
	"Category type",
	"Modality",
	"Match scope",
	"Min score (trials)",
	"Field weights (trials)",
	"Trials (this subject)",
	"Last synced",
]

CATEGORY_MATCH_PROSE = (
	"Each row is one category assigned to one subject. Automatic categories are populated "
	"by the rebuild_categories command: each search term is matched case-insensitively as a "
	"whole word against the fields listed in \"Field weights (trials)\" (only those fields "
	"are searched, per the category's match scope); every matching field adds its weight, "
	"plus a flat 2 points per unique matched term, and a trial is assigned once the total "
	"reaches the \"Min score (trials)\" threshold. Manual categories are curated by hand and "
	"ignore the term list."
)

# Column widths for data sheets
_WIDE_COLS = {
	"title",
	"scientific_title",
	"summary",
	"ctg_detailed_description",
	"inclusion_criteria",
	"exclusion_criteria",
	"intervention",
	"primary_outcome",
	"secondary_outcome",
	"study_design",
	"condition",
	"results_ipd_description",
	"therapeutic_areas",
	"country_status",
	"countries_by_source",
	"trial_countries",
	"articles",
}
_URL_COLS = {"link", "results_url_link", "identifiers_json"}


def _sanitise_sheet_name(name, used):
	"""Return a valid, unique Excel sheet name (≤31 chars, no illegal characters)."""
	name = re.sub(r"[\\/*?\[\]:]", "_", name)[:31]
	base, suffix = name, 1
	while name in used:
		suffix += 1
		name = base[:28] + f"_{suffix}"
	used.add(name)
	return name


def _cell_value(value):
	"""Convert a Python value to an Excel-safe type (None → empty string)."""
	if value is None:
		return ""
	if isinstance(value, datetime):
		if value.tzinfo is not None:
			value = value.astimezone(dt_timezone.utc).replace(tzinfo=None)
		return value
	if isinstance(value, date):
		return value
	if isinstance(value, dict):
		return json.dumps(value, ensure_ascii=False)
	if isinstance(value, list):
		# Lists of scalars (e.g. regions_normalized) render as a "; "-joined string;
		# lists containing dicts/lists fall back to JSON for fidelity.
		if all(not isinstance(v, (dict, list)) for v in value):
			return "; ".join("" if v is None else str(v) for v in value)
		return json.dumps(value, ensure_ascii=False)
	if isinstance(value, bool):
		return value
	if isinstance(value, (int, float)):
		return value
	return str(value)


def _format_trial_countries(trial):
	"""Render a trial's normalized TrialCountry rows as one readable cell.

	Format per country: "Germany [DE] (recruiting; 2024-07-19; src: ctgov+ctis)".
	The status/date/sources clause is omitted when empty. Uses the prefetched
	``trial_countries`` cache (sorted in Python — no extra query).
	"""
	rows = sorted(trial.trial_countries.all(), key=lambda tc: str(tc.country.code))
	parts = []
	for tc in rows:
		bits = []
		status = tc.status_raw or tc.status
		if status:
			bits.append(str(status))
		if tc.decision_date:
			bits.append(str(tc.decision_date))
		if tc.sources:
			bits.append("src: " + "+".join(tc.sources))
		label = f"{tc.country.name} [{tc.country.code}]"
		if bits:
			label += " (" + "; ".join(bits) + ")"
		parts.append(label)
	return "; ".join(parts)


def _apply_header(ws, columns, row=1):
	"""Write a bold, coloured header row and return the cell count."""
	hdr_font = Font(bold=True, color="FFFFFF")
	hdr_fill = PatternFill(fill_type="solid", fgColor="2F4F8F")
	hdr_align = Alignment(vertical="center")
	for col_idx, name in enumerate(columns, 1):
		cell = ws.cell(row=row, column=col_idx, value=name)
		cell.font = hdr_font
		cell.fill = hdr_fill
		cell.alignment = hdr_align
	return len(columns)


def _excel_text(value, limit=32767):
	"""Truncate a string to Excel's per-cell character limit."""
	if not value:
		return value
	text = str(value)
	if len(text) > limit:
		text = text[: limit - 1] + "…"
	return text


_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@")


def _write_safe_text_cell(ws, row, col, value, wrap=False):
	"""Write a long/user-authored text value, truncated and defused against formula injection.

	A leading apostrophe is the standard mitigation (OWASP CSV injection guidance):
	it forces the cell to render as literal text in Excel/LibreOffice/Sheets even
	when the value starts with =, +, -, or @. Relying on openpyxl's data_type alone
	is not enough — several of those trigger characters (+, -, @) are never
	auto-detected as formulas by openpyxl in the first place, so its cell.data_type
	would stay "s" regardless, while spreadsheet applications still treat a leading
	=, +, -, or @ as a formula cue on their own.
	"""
	text = _excel_text(value)
	if isinstance(text, str) and text.startswith(_FORMULA_TRIGGER_CHARS):
		text = "'" + text
	cell = ws.cell(row=row, column=col, value=text)
	cell.data_type = "s"
	if wrap:
		cell.alignment = Alignment(wrap_text=True)
	return cell


def _set_column_widths(ws, columns):
	for col_idx, name in enumerate(columns, 1):
		letter = get_column_letter(col_idx)
		if name in _WIDE_COLS:
			ws.column_dimensions[letter].width = 50
		elif name in _URL_COLS:
			ws.column_dimensions[letter].width = 40
		elif name.startswith("id_"):
			ws.column_dimensions[letter].width = 20
		else:
			ws.column_dimensions[letter].width = max(12, min(40, len(name) + 4))


def _build_scalar_columns():
	"""
	Return an ordered list of scalar Trials column names for export.
	Known columns follow IDENTITY_COLS + SCALAR_ORDER; unrecognised columns are appended.
	Excludes: GeneratedField columns, m2m/relation fields, and 'identifiers' (expanded
	separately into id_* columns and identifiers_json).
	"""
	known_order = IDENTITY_COLS + SCALAR_ORDER
	# 'identifiers' is handled separately; EXCLUDED_* are never exported
	known_set = set(known_order) | {"identifiers"} | EXCLUDED_SCALARS | EXCLUDED_M2M
	unknown = []
	for f in Trials._meta.get_fields():
		if f.is_relation:
			continue
		if isinstance(f, GeneratedField):
			continue
		if f.name in known_set:
			continue
		unknown.append(f.name)
	return known_order + sorted(unknown)


def _parse_help_text(text):
	"""
	Split a help_text string into (description, source_registries_string).
	Looks for a trailing 'Sources?: …' clause.
	"""
	match = re.search(r"\s+Sources?:\s*(.+?)\.?\s*$", text, re.IGNORECASE)
	if match:
		return text[: match.start()].strip(), match.group(1).strip().rstrip(".")
	return text.strip(), ""


def _sources_for(col_name, admin_labels, admin_help, model_help):
	"""Return (label, description, source_str) for one exported column."""
	if col_name.startswith("id_"):
		key = col_name[3:]
		return (
			f"Identifier: {key.upper()}",
			f'Registry identifier key "{key}" extracted from the identifiers JSON.',
			"WHO ICTRP, ClinicalTrials.gov, EU CTIS",
		)
	if col_name in EXTRA_GLOSSARY:
		return EXTRA_GLOSSARY[col_name]
	if col_name in admin_help:
		label = admin_labels.get(col_name, col_name)
		desc, sources = _parse_help_text(admin_help[col_name])
		return label, desc, sources
	if col_name in model_help:
		label = admin_labels.get(col_name, col_name)
		desc, sources = _parse_help_text(model_help[col_name])
		return label, desc, sources
	return admin_labels.get(col_name, col_name), "", ""


def _build_glossary_sheet(wb, all_data_cols):
	"""Add a Glossary sheet — one row per exported column."""
	from gregory.admin import TrialAdminForm

	admin_labels = TrialAdminForm.Meta.labels
	admin_help = TrialAdminForm.Meta.help_texts
	model_help = {
		f.name: f.help_text
		for f in Trials._meta.get_fields()
		if not f.is_relation and hasattr(f, "help_text") and f.help_text
	}

	ws = wb.create_sheet(title="Glossary")
	headers = ["Field", "Label", "Description", "Source registries"]
	_apply_header(ws, headers)
	ws.freeze_panes = "A2"
	ws.column_dimensions["A"].width = 32
	ws.column_dimensions["B"].width = 32
	ws.column_dimensions["C"].width = 65
	ws.column_dimensions["D"].width = 32

	for row_idx, col_name in enumerate(all_data_cols, 2):
		label, desc, sources = _sources_for(
			col_name, admin_labels, admin_help, model_help
		)
		ws.cell(row=row_idx, column=1, value=col_name)
		ws.cell(row=row_idx, column=2, value=label)
		cell_desc = ws.cell(row=row_idx, column=3, value=desc)
		cell_desc.alignment = Alignment(wrap_text=True)
		ws.cell(row=row_idx, column=4, value=sources)


def _build_registries_sheet(wb, all_data_cols):
	"""Add a Registries sheet with a prose overview, registry table, and field matrix."""
	from gregory.admin import TrialAdminForm

	admin_labels = TrialAdminForm.Meta.labels
	admin_help = TrialAdminForm.Meta.help_texts
	model_help = {
		f.name: f.help_text
		for f in Trials._meta.get_fields()
		if not f.is_relation and hasattr(f, "help_text") and f.help_text
	}

	ws = wb.create_sheet(title="Registries")
	hdr_font = Font(bold=True, color="FFFFFF")
	hdr_fill = PatternFill(fill_type="solid", fgColor="2F4F8F")
	hdr_align = Alignment(vertical="center")
	section_font = Font(bold=True, size=13)

	# --- Part A: prose overview ---
	ws.cell(row=1, column=1, value="Registry overview").font = section_font
	prose_cell = ws.cell(row=2, column=1, value=MERGE_PROSE)
	prose_cell.alignment = Alignment(wrap_text=True)
	ws.merge_cells("A2:D2")
	ws.row_dimensions[2].height = 90

	# Overview table header (row 4)
	for col_idx, h in enumerate(
		["Registry", "Identifier key(s)", "What it is", "Notes on coverage"], 1
	):
		c = ws.cell(row=4, column=col_idx, value=h)
		c.font = hdr_font
		c.fill = hdr_fill
		c.alignment = hdr_align

	for row_offset, (reg_name, id_keys_str, what, notes) in enumerate(
		REGISTRIES_OVERVIEW, 5
	):
		ws.cell(row=row_offset, column=1, value=reg_name)
		ws.cell(row=row_offset, column=2, value=id_keys_str)
		ws.cell(row=row_offset, column=3, value=what)
		ws.cell(row=row_offset, column=4, value=notes).alignment = Alignment(
			wrap_text=True
		)

	# --- Part B: field-by-registry matrix ---
	matrix_start = 10
	ws.cell(
		row=matrix_start, column=1, value="Field coverage by registry"
	).font = section_font

	for col_idx, h in enumerate(["Field", "Label"] + REGISTRY_NAMES, 1):
		c = ws.cell(row=matrix_start + 1, column=col_idx, value=h)
		c.font = hdr_font
		c.fill = hdr_fill
		c.alignment = hdr_align

	for row_offset, col_name in enumerate(all_data_cols, matrix_start + 2):
		label, _, sources_str = _sources_for(
			col_name, admin_labels, admin_help, model_help
		)
		ws.cell(row=row_offset, column=1, value=col_name)
		ws.cell(row=row_offset, column=2, value=label)
		for reg_col, reg_name in enumerate(REGISTRY_NAMES, 3):
			# Normalise for matching: "EU CTIS" matches "EU CTIS" or "EU Clinical"
			src_lower = sources_str.lower()
			reg_lower = reg_name.lower()
			tick = "✓" if reg_lower in src_lower else ""
			if not tick and reg_name == "EU CTIS" and "eu ctis" not in src_lower:
				if "eu clinical" in src_lower:
					tick = "✓"
			ws.cell(row=row_offset, column=reg_col, value=tick)

	# Column widths
	for letter, width in zip("ABCDE", [32, 32, 22, 22, 22]):
		ws.column_dimensions[letter].width = width


def _category_rows(subjects):
	"""Return one row per (subject, category) pair, ordered by subject then category name.

	Categories with no subject never appear here: filtering by `subjects=subject`
	only matches categories reachable from an exported subject.
	"""
	rows = []
	for subject in subjects:
		cats = list(
			TeamCategory.objects.filter(subjects=subject)
			.select_related("team")
			.order_by(Lower("category_name"))
		)
		if not cats:
			continue
		counts = {
			r["team_categories"]: r["n"]
			for r in (
				Trials.objects.filter(
					subjects=subject, team_categories__in=[c.pk for c in cats]
				)
				.values("team_categories")
				.annotate(n=Count("pk", distinct=True))
			)
		}
		for cat in cats:
			weights = cat.get_scored_fields("trial")
			weights_str = "; ".join(f"{f}:{w}" for f, w in weights.items() if w)
			rows.append(
				(
					subject.subject_name,
					cat.team.name,
					cat.category_name,
					cat.category_slug or "",
					cat.category_description or "",
					"; ".join(cat.category_terms or []),
					len(cat.category_terms or []),
					cat.get_category_type_display(),
					cat.get_modality_display() if cat.modality else "",
					cat.get_match_scope_display(),
					cat.match_min_score_trials,
					weights_str,
					counts.get(cat.pk, 0),
					_cell_value(cat.last_synced_at),
				)
			)
	return rows


def _build_categories_sheet(wb, subjects):
	"""Add a Categories sheet — one row per (subject, category) pair."""
	ws = wb.create_sheet(title="Categories")
	section_font = Font(bold=True, size=13)

	ws.cell(
		row=1, column=1, value="Categories and their search terms"
	).font = section_font
	prose_cell = ws.cell(row=2, column=1, value=CATEGORY_MATCH_PROSE)
	prose_cell.alignment = Alignment(wrap_text=True)
	last_col_letter = get_column_letter(len(CATEGORY_COLUMNS))
	ws.merge_cells(f"A2:{last_col_letter}2")
	ws.row_dimensions[2].height = 90

	_apply_header(ws, CATEGORY_COLUMNS, row=4)
	ws.freeze_panes = "A5"
	ws.auto_filter.ref = f"A4:{last_col_letter}4"

	rows = _category_rows(subjects)
	desc_col = CATEGORY_COLUMNS.index("Description") + 1
	terms_col = CATEGORY_COLUMNS.index("Search terms") + 1

	if not rows:
		ws.cell(row=5, column=1, value="No categories found.")
	else:
		for row_idx, row_data in enumerate(rows, 5):
			for col_idx, value in enumerate(row_data, 1):
				if col_idx in (desc_col, terms_col):
					_write_safe_text_cell(ws, row_idx, col_idx, value, wrap=True)
				else:
					ws.cell(row=row_idx, column=col_idx, value=value)

	wide_widths = {"Description": 65, "Search terms": 65, "Field weights (trials)": 40}
	for col_idx, name in enumerate(CATEGORY_COLUMNS, 1):
		letter = get_column_letter(col_idx)
		ws.column_dimensions[letter].width = wide_widths.get(
			name, max(12, min(32, len(name) + 4))
		)


GENERATED_BY_PROSE = "GregoryAI — https://github.com/brunoamaral/gregory-ai"

WHATS_INSIDE_FIXED_ROWS = [
	("Categories", "How trials are tagged, with each category's search terms."),
	("Glossary", "What every column in the subject sheets means."),
	(
		"Registries",
		"Which clinical-trial registries the data came from, and how records from "
		"several registries are merged.",
	),
]


@dataclass
class SiteAttribution:
	"""Which Site (and its CustomSetting) an export is attributed to."""

	site: object = None
	custom_setting: object = None
	other_sites: list = field(default_factory=list)


def _resolve_site_for_team(team):
	"""Resolve a single team's Site: team.site, else the org's default/first site, else the current site."""
	if team is None:
		return None
	if team.site_id:
		return team.site
	if team.organization_id:
		org_site = (
			OrganizationSite.objects.filter(organization_id=team.organization_id)
			.order_by("-is_default", "id")
			.select_related("site")
			.first()
		)
		if org_site:
			return org_site.site
	try:
		return Site.objects.get_current()
	except Site.DoesNotExist:
		return None


def _resolve_default_site(subjects):
	"""Resolve the Site to attribute an export to, from its subjects' teams.

	Returns (site_or_None, other_sites) where other_sites lists every other site
	found (sorted by pk) when the export spans more than one.
	"""
	counts = Counter()
	by_pk = {}
	for subject in subjects:
		site = _resolve_site_for_team(subject.team)
		if site is None:
			continue
		counts[site.pk] += 1
		by_pk[site.pk] = site

	if not counts:
		return None, []
	if len(counts) == 1:
		(only_pk,) = counts.keys()
		return by_pk[only_pk], []

	best_pk = max(counts, key=lambda pk: (counts[pk], -pk))
	others = [by_pk[pk] for pk in sorted(counts) if pk != best_pk]
	return by_pk[best_pk], others


def _resolve_explicit_site(value):
	"""Resolve --site by numeric ID or domain (case-insensitive). Raises CommandError if not found."""
	value = value.strip()
	if value.isdigit():
		site = Site.objects.filter(pk=int(value)).first()
	else:
		site = Site.objects.filter(domain__iexact=value).first()
	if site is None:
		listing = ", ".join(f"{s.pk} ({s.domain})" for s in Site.objects.order_by("pk"))
		raise CommandError(f"Site not found: {value!r}. Valid sites: {listing}")
	return site


def _resolve_site_attribution(subjects, explicit_site_value):
	"""Resolve the SiteAttribution for an export, from --site or the subjects' teams."""
	if explicit_site_value:
		site = _resolve_explicit_site(explicit_site_value)
		other_sites = []
	else:
		site, other_sites = _resolve_default_site(subjects)

	custom_setting = None
	if site is not None:
		custom_setting = (
			CustomSetting.objects.filter(site=site).order_by("setting_id").first()
		)
	return SiteAttribution(site=site, custom_setting=custom_setting, other_sites=other_sites)


def _build_about_sheet(ws, attribution, sheet_entries, options_summary):
	"""Populate the 'About this file' sheet: Source, This file, What's in this workbook."""
	section_fill = PatternFill(fill_type="solid", fgColor="2F4F8F")
	section_font = Font(bold=True, size=13, color="FFFFFF")
	label_font = Font(bold=True)

	ws.column_dimensions["A"].width = 28
	ws.column_dimensions["B"].width = 90

	row = 1

	def write_section_header(text):
		nonlocal row
		cell = ws.cell(row=row, column=1, value=text)
		cell.font = section_font
		cell.fill = section_fill
		ws.cell(row=row, column=2).fill = section_fill
		ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
		row += 1

	def write_row(label, value, wrap=True):
		nonlocal row
		if not value:
			return
		# Most labels are our own fixed strings, but Section 3's are subject
		# names, so route every label through the same defusing as the value.
		_write_safe_text_cell(ws, row, 1, label, wrap=False)
		ws.cell(row=row, column=1).font = label_font
		_write_safe_text_cell(ws, row, 2, value, wrap=wrap)
		row += 1

	def write_note(text):
		nonlocal row
		_write_safe_text_cell(ws, row, 1, text, wrap=True)
		ws.cell(row=row, column=1).font = Font(italic=True)
		ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
		row += 1

	site = attribution.site
	cs = attribution.custom_setting

	write_section_header("Source")
	if site is None:
		write_note(
			"No site is configured for the exported subjects' teams. Set a site on "
			"the team, or pass --site."
		)
	else:
		title = (cs.title if cs and cs.title else None) or site.name or site.domain
		write_row("Published by", title)

		org_site = (
			OrganizationSite.objects.filter(site=site)
			.select_related("organization")
			.first()
		)
		org_name = org_site.organization.name if org_site else ""
		if org_name and org_name != title:
			write_row("Organisation", org_name)

		if cs:
			write_row("About this project", cs.description)
			write_row("Website", cs.website_url or f"https://{site.domain}")
			write_row("About page", cs.about_url)
			write_row("Contact", cs.contact_url)
			write_row("Contact email", cs.contact_email or cs.admin_email)
			if cs.api_domain:
				write_row("API", f"https://{cs.api_domain}")
			write_row("Source code", cs.github_url)
			social = "; ".join(
				filter(
					None,
					[
						f"Mastodon: {cs.mastodon_url}" if cs.mastodon_url else "",
						f"Bluesky: {cs.bluesky_url}" if cs.bluesky_url else "",
					],
				)
			)
			write_row("Mastodon / Bluesky", social)
			write_row("Privacy policy", cs.privacy_policy_url)
		else:
			write_row("Website", f"https://{site.domain}")

	if attribution.other_sites and site is not None:
		all_sites_sorted = sorted([site] + attribution.other_sites, key=lambda s: s.pk)
		domains = ", ".join(s.domain for s in all_sites_sorted if s.pk != site.pk)
		write_note(
			f"This workbook also contains subjects published by {domains}. Pass "
			"--site to attribute it explicitly."
		)

	write_section_header("This file")
	generated = datetime.now(dt_timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
	write_row("Generated", generated)
	write_row("Generated by", GENERATED_BY_PROSE)
	write_row("Export options", options_summary)

	data_license = cs.data_license if cs else ""
	data_license_url = cs.data_license_url if cs else ""
	write_row("Data licence", data_license)
	write_row("Licence URL", data_license_url)

	citation = cs.citation if cs else ""
	if not citation:
		title_for_citation = (
			(cs.title if cs and cs.title else None)
			or (site.name or site.domain if site else "")
			or "GregoryAI export"
		)
		website_for_citation = (cs.website_url if cs else "") or (
			f"https://{site.domain}" if site else ""
		)
		date_str = date.today().strftime("%Y-%m-%d")
		citation = f"{title_for_citation}. Clinical trials export, {date_str}."
		if website_for_citation:
			citation += f" {website_for_citation}"
	write_row("How to cite", citation)

	write_section_header("What's in this workbook")
	for sheet_name, subject, count in sheet_entries:
		value = f'{count} clinical trial(s) for the research subject "{subject.subject_name}"'
		if subject.description:
			value += f" — {subject.description}"
		write_row(sheet_name, value)
	for label, value in WHATS_INSIDE_FIXED_ROWS:
		write_row(label, value)


class Command(BaseCommand):
	help = "Export clinical-trial data to an XLSX workbook, one sheet per subject."

	def add_arguments(self, parser):
		parser.add_argument(
			"--subjects",
			type=str,
			default="",
			help="Comma-separated subject IDs to export.",
		)
		parser.add_argument(
			"--all-subjects",
			action="store_true",
			default=False,
			help="Export every subject (one sheet each).",
		)
		parser.add_argument(
			"--output",
			type=str,
			default="",
			help="Output file path (default: trials_export_YYYYMMDD.xlsx in current directory).",
		)
		parser.add_argument(
			"--team",
			type=int,
			default=None,
			help="Optional team ID; filters which subjects are exported.",
		)
		parser.add_argument(
			"--site",
			type=str,
			default="",
			help="Site ID or domain to attribute this export to. Defaults to the site "
			"resolved from the exported subjects' teams.",
		)

	def handle(self, *args, **options):
		# --- Resolve subjects ---
		if options["all_subjects"]:
			qs = Subject.objects.all()
			if options["team"]:
				qs = qs.filter(team_id=options["team"])
			# select_related avoids one query per subject in _resolve_default_site,
			# which walks subject.team and team.site for every exported subject.
			subjects = list(
				qs.select_related("team", "team__site").order_by("subject_name")
			)
		else:
			raw = options["subjects"].strip()
			if not raw:
				raise CommandError("Provide --subjects <ids> or --all-subjects.")
			try:
				ids = [int(x.strip()) for x in raw.split(",") if x.strip()]
			except ValueError:
				raise CommandError("--subjects must be comma-separated integers.")
			subject_qs = Subject.objects.all()
			if options["team"]:
				subject_qs = subject_qs.filter(team_id=options["team"])
			valid_ids = set(subject_qs.filter(pk__in=ids).values_list("pk", flat=True))
			missing = set(ids) - valid_ids
			if missing:
				all_subs = subject_qs.values_list("pk", "subject_name")
				valid_list = ", ".join(f"{pk} ({name})" for pk, name in all_subs)
				raise CommandError(
					f"Subject ID(s) not found: {sorted(missing)}. Valid IDs: {valid_list}"
				)
			subjects = list(
				subject_qs.filter(pk__in=ids)
				.select_related("team", "team__site")
				.order_by("subject_name")
			)

		if not subjects:
			raise CommandError("No subjects found.")

		# --- Resolve site attribution ---
		attribution = _resolve_site_attribution(subjects, options["site"])
		if attribution.other_sites:
			all_sites_sorted = sorted(
				[attribution.site] + attribution.other_sites, key=lambda s: s.pk
			)
			domains = ", ".join(s.domain for s in all_sites_sorted)
			self.stdout.write(
				self.style.WARNING(
					f"Exported subjects span multiple sites ({domains}); attributing "
					f"this export to {attribution.site.domain} (most subjects). Pass "
					"--site to be explicit."
				)
			)

		output_path = (
			options["output"]
			or f"trials_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
		)

		# --- Build column plan ---
		scalar_cols = _build_scalar_columns()
		remaining_scalars = [
			c for c in scalar_cols if c not in set(IDENTITY_COLS) and c != "identifiers"
		]

		# Discover all identifier keys across every exported subject (preserves first-seen order)
		all_subject_ids = [s.pk for s in subjects]
		id_key_order, seen_keys = [], set()
		for identifiers in (
			Trials.objects.filter(subjects__in=all_subject_ids)
			.exclude(identifiers=None)
			.values_list("identifiers", flat=True)
		):
			if isinstance(identifiers, dict):
				for k in identifiers:
					if k not in seen_keys:
						id_key_order.append(k)
						seen_keys.add(k)

		id_cols = [f"id_{k}" for k in id_key_order]
		all_data_cols = (
			IDENTITY_COLS
			+ id_cols
			+ ["identifiers_json"]
			+ remaining_scalars
			+ RELATION_COLS
		)

		self.stdout.write(f"Exporting {len(subjects)} subject(s) → {output_path}")

		# --- Build workbook ---
		wb = Workbook()
		wb.remove(wb.active)  # remove the default blank sheet
		used_sheet_names: set = {"About this file", "Categories", "Glossary", "Registries"}
		ws_about = wb.create_sheet(title="About this file")
		sheet_entries = []  # (sheet_name, subject, count)

		for subject in subjects:
			sheet_name = _sanitise_sheet_name(subject.subject_name, used_sheet_names)
			ws = wb.create_sheet(title=sheet_name)

			qs = (
				Trials.objects.filter(subjects=subject)
				.distinct()
				.order_by("-discovery_date")
				.select_related("primary_sponsor_normalized")
				.prefetch_related(
					"subjects",
					"teams",
					"sources",
					"team_categories",
					"article_references__article",
					"trial_countries",
				)
			)

			count = qs.count()
			self.stdout.write(f'  Sheet "{sheet_name}": {count} trial(s)')

			_apply_header(ws, all_data_cols)
			ws.freeze_panes = "A2"
			ws.auto_filter.ref = f"A1:{get_column_letter(len(all_data_cols))}1"

			if count == 0:
				ws.cell(row=2, column=1, value="No trials found for this subject.")
			else:
				scalar_set = set(IDENTITY_COLS) | set(remaining_scalars)
				for row_idx, trial in enumerate(qs, 2):
					identifiers = trial.identifiers or {}
					row_data = []
					for col_name in all_data_cols:
						if col_name in scalar_set:
							row_data.append(_cell_value(getattr(trial, col_name, None)))
						elif col_name.startswith("id_"):
							row_data.append(_cell_value(identifiers.get(col_name[3:])))
						elif col_name == "identifiers_json":
							row_data.append(
								json.dumps(identifiers, ensure_ascii=False)
								if identifiers
								else ""
							)
						elif col_name == "subjects":
							row_data.append(
								"; ".join(s.subject_name for s in trial.subjects.all())
							)
						elif col_name == "teams":
							row_data.append(
								"; ".join(t.name for t in trial.teams.all())
							)
						elif col_name == "sources":
							row_data.append(
								"; ".join(src.name or "" for src in trial.sources.all())
							)
						elif col_name == "team_categories":
							row_data.append(
								"; ".join(
									tc.category_name
									for tc in trial.team_categories.all()
								)
							)
						elif col_name == "articles":
							refs = list(trial.article_references.all())
							if refs:
								links = "; ".join(r.article.link for r in refs)
								row_data.append(f"{len(refs)}: {links}")
							else:
								row_data.append("")
						elif col_name == "trial_countries":
							row_data.append(_format_trial_countries(trial))
						elif col_name == "sponsor_id":
							sponsor = trial.primary_sponsor_normalized
							row_data.append(sponsor.pk if sponsor else "")
						elif col_name == "sponsor_slug":
							sponsor = trial.primary_sponsor_normalized
							row_data.append(sponsor.slug if sponsor else "")
						elif col_name == "primary_sponsor_normalized":
							sponsor = trial.primary_sponsor_normalized
							row_data.append(sponsor.name if sponsor else "")
						elif col_name == "sponsor_type_normalized":
							sponsor = trial.primary_sponsor_normalized
							row_data.append(
								(sponsor.sponsor_type or "") if sponsor else ""
							)
						elif col_name == "sponsor_type_source":
							sponsor = trial.primary_sponsor_normalized
							row_data.append(
								(sponsor.sponsor_type_source or "") if sponsor else ""
							)
						else:
							row_data.append("")

					for col_idx, value in enumerate(row_data, 1):
						ws.cell(row=row_idx, column=col_idx, value=value)

			_set_column_widths(ws, all_data_cols)
			sheet_entries.append((sheet_name, subject, count))

		_build_categories_sheet(wb, subjects)
		_build_glossary_sheet(wb, all_data_cols)
		_build_registries_sheet(wb, all_data_cols)

		options_summary_parts = [
			"subjects: " + ", ".join(s.subject_name for s in subjects)
		]
		if options["team"]:
			options_summary_parts.append(f"team: {options['team']}")
		if attribution.site:
			options_summary_parts.append(f"site: {attribution.site.domain}")
		options_summary = "; ".join(options_summary_parts)

		_build_about_sheet(ws_about, attribution, sheet_entries, options_summary)

		wb.save(output_path)
		self.stdout.write(self.style.SUCCESS(f"Saved: {output_path}"))
