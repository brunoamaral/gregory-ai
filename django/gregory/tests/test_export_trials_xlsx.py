"""
Tests for the export_trials_xlsx management command.
Run in the gregory Docker container:
    python manage.py test gregory.tests.test_export_trials_xlsx
"""

import os
import tempfile
from io import StringIO

from django.contrib.sites.models import Site
from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase
from organizations.models import Organization

import openpyxl

from gregory.models import (
	Articles,
	ArticleTrialReference,
	CategoryModality,
	CategoryType,
	OrganizationSite,
	Sources,
	Subject,
	Team,
	TeamCategory,
	Trials,
	TrialCategoryAssignment,
	TrialCountry,
)
from gregory.management.commands.export_trials_xlsx import (
	CATEGORY_COLUMNS,
	IDENTITY_COLS,
	RELATION_COLS,
	REGISTRY_NAMES,
	_build_scalar_columns,
	_sanitise_sheet_name,
)
from sitesettings.models import CustomSetting


class ExportTrialsXlsxTests(TestCase):
	def setUp(self):
		self.org = Organization.objects.create(name="Test Org", slug="test-org")
		self.team = Team.objects.create(
			name="Team A", organization=self.org, slug="team-a"
		)

		self.subject_ms = Subject.objects.create(
			subject_name="Multiple Sclerosis", subject_slug="ms", team=self.team
		)
		self.subject_cancer = Subject.objects.create(
			subject_name="Cancer", subject_slug="cancer", team=self.team
		)

		self.source_ctg = Sources.objects.create(
			name="ClinicalTrials.gov", source_for="trials"
		)
		self.source_who = Sources.objects.create(name="WHO ICTRP", source_for="trials")

		# Trial assigned to MS subject — NCT identifier
		self.trial_ms = Trials.objects.create(
			title="A randomised trial of natalizumab in MS",
			link="https://clinicaltrials.gov/ct2/show/NCT00111111",
			identifiers={"nct": "NCT00111111"},
			recruitment_status="Completed",
			phase="Phase 3",
		)
		self.trial_ms.subjects.add(self.subject_ms)
		self.trial_ms.sources.add(self.source_ctg)

		# Trial assigned to MS subject — EUCTR identifier
		self.trial_ms2 = Trials.objects.create(
			title="An EU trial on ocrelizumab in MS",
			link="https://www.clinicaltrialsregister.eu/ctr-search/trial/2016-001005-36/ES",
			identifiers={"euctr": "2016-001005-36"},
			recruitment_status="Ongoing",
			therapeutic_areas="Neurology",
		)
		self.trial_ms2.subjects.add(self.subject_ms)
		self.trial_ms2.sources.add(self.source_who)

		# Trial assigned to Cancer subject only
		self.trial_cancer = Trials.objects.create(
			title="A trial for pancreatic cancer",
			link="https://clinicaltrials.gov/ct2/show/NCT00999999",
			identifiers={"nct": "NCT00999999"},
			recruitment_status="Recruiting",
		)
		self.trial_cancer.subjects.add(self.subject_cancer)
		self.trial_cancer.sources.add(self.source_ctg)

		# Article referencing the MS trial
		self.article = Articles.objects.create(
			title="Natalizumab review",
			link="https://pubmed.ncbi.nlm.nih.gov/12345678/",
		)
		ArticleTrialReference.objects.create(
			article=self.article,
			trial=self.trial_ms,
			identifier_type="nct",
			identifier_value="NCT00111111",
		)

		# Categories exercising the Categories sheet
		self.cat_natalizumab = TeamCategory.objects.create(
			team=self.team,
			category_name="Natalizumab",
			category_description="Trials studying natalizumab (Tysabri) in MS.",
			category_terms=["natalizumab", "tysabri"],
			category_type=CategoryType.AUTOMATIC,
			modality=CategoryModality.BIOLOGIC_ANTIBODY,
		)
		self.cat_natalizumab.subjects.add(self.subject_ms)
		TrialCategoryAssignment.objects.create(
			trials=self.trial_ms, teamcategory=self.cat_natalizumab
		)

		self.cat_manual = TeamCategory.objects.create(
			team=self.team,
			category_name="Manual watchlist",
			category_type=CategoryType.MANUAL,
			category_terms=[],
		)
		self.cat_manual.subjects.add(self.subject_ms)

		self.cat_shared = TeamCategory.objects.create(
			team=self.team,
			category_name="Shared category",
			category_terms=["shared-term"],
		)
		self.cat_shared.subjects.add(self.subject_ms, self.subject_cancer)

		self.cat_orphan = TeamCategory.objects.create(
			team=self.team,
			category_name="Orphan category",
			category_terms=["orphan-term"],
		)

	def _export(self, **kwargs):
		"""Run the command into a temp file; return (path, workbook)."""
		fd, path = tempfile.mkstemp(suffix=".xlsx")
		os.close(fd)
		call_command("export_trials_xlsx", output=path, **kwargs)
		wb = openpyxl.load_workbook(path)
		return path, wb

	def tearDown(self):
		pass  # temp files cleaned up by individual tests

	# ------------------------------------------------------------------
	# Sheet structure
	# ------------------------------------------------------------------

	def test_one_sheet_per_subject_plus_glossary_and_registries(self):
		path, wb = self._export(
			subjects=f"{self.subject_ms.pk},{self.subject_cancer.pk}"
		)
		try:
			sheet_names = wb.sheetnames
			self.assertIn("Multiple Sclerosis", sheet_names)
			self.assertIn("Cancer", sheet_names)
			self.assertIn("Glossary", sheet_names)
			self.assertIn("Registries", sheet_names)
		finally:
			os.unlink(path)

	def test_all_subjects_flag(self):
		path, wb = self._export(all_subjects=True)
		try:
			sheet_names = wb.sheetnames
			self.assertIn("Multiple Sclerosis", sheet_names)
			self.assertIn("Cancer", sheet_names)
		finally:
			os.unlink(path)

	def test_subject_not_found_raises(self):
		with self.assertRaises(CommandError):
			call_command(
				"export_trials_xlsx", subjects="99999", output="/tmp/noop.xlsx"
			)

	def test_no_args_raises(self):
		with self.assertRaises(CommandError):
			call_command("export_trials_xlsx", output="/tmp/noop.xlsx")

	# ------------------------------------------------------------------
	# Data sheet headers
	# ------------------------------------------------------------------

	def test_identity_columns_present(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Multiple Sclerosis"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			for col in IDENTITY_COLS:
				self.assertIn(
					col, headers, f'Identity column "{col}" missing from header row'
				)
		finally:
			os.unlink(path)

	def test_id_columns_created_from_identifier_keys(self):
		"""id_nct and id_euctr must appear as explicit columns."""
		path, wb = self._export(
			subjects=f"{self.subject_ms.pk},{self.subject_cancer.pk}"
		)
		try:
			ws = wb["Multiple Sclerosis"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			self.assertIn("id_nct", headers)
			self.assertIn("id_euctr", headers)
			self.assertIn("identifiers_json", headers)
		finally:
			os.unlink(path)

	def test_relation_columns_present(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Multiple Sclerosis"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			for col in RELATION_COLS:
				self.assertIn(
					col, headers, f'Relation column "{col}" missing from header row'
				)
		finally:
			os.unlink(path)

	# ------------------------------------------------------------------
	# Data values
	# ------------------------------------------------------------------

	def test_trial_data_written_correctly(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Multiple Sclerosis"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			# Find the title column
			title_col = headers.index("title") + 1
			titles = [
				ws.cell(row=r, column=title_col).value for r in range(2, ws.max_row + 1)
			]
			self.assertIn("A randomised trial of natalizumab in MS", titles)
			self.assertIn("An EU trial on ocrelizumab in MS", titles)
		finally:
			os.unlink(path)

	def test_identifier_expansion(self):
		"""id_nct column should hold the NCT number; id_euctr should hold the EUCTR number."""
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Multiple Sclerosis"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			nct_col = headers.index("id_nct") + 1
			euctr_col = headers.index("id_euctr") + 1
			title_col = headers.index("title") + 1

			nct_values = {}
			euctr_values = {}
			for r in range(2, ws.max_row + 1):
				title = ws.cell(row=r, column=title_col).value
				nct_values[title] = ws.cell(row=r, column=nct_col).value
				euctr_values[title] = ws.cell(row=r, column=euctr_col).value

			self.assertEqual(
				nct_values.get("A randomised trial of natalizumab in MS"), "NCT00111111"
			)
			self.assertEqual(
				euctr_values.get("An EU trial on ocrelizumab in MS"), "2016-001005-36"
			)
		finally:
			os.unlink(path)

	def test_articles_column_shows_link(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Multiple Sclerosis"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			articles_col = headers.index("articles") + 1
			title_col = headers.index("title") + 1

			for r in range(2, ws.max_row + 1):
				if (
					ws.cell(row=r, column=title_col).value
					== "A randomised trial of natalizumab in MS"
				):
					cell_val = ws.cell(row=r, column=articles_col).value
					self.assertIn("pubmed.ncbi.nlm.nih.gov", cell_val or "")
					break
			else:
				self.fail("MS trial row not found")
		finally:
			os.unlink(path)

	def test_empty_subject_still_has_header(self):
		"""A subject with no trials must still produce a headed (non-crashing) sheet."""
		empty_subject = Subject.objects.create(
			subject_name="Empty Subject", subject_slug="empty-subject", team=self.team
		)
		path, wb = self._export(subjects=str(empty_subject.pk))
		try:
			ws = wb["Empty Subject"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			self.assertIn("trial_id", headers)
		finally:
			os.unlink(path)

	# ------------------------------------------------------------------
	# Normalized country columns
	# ------------------------------------------------------------------

	def _cell_for_title(self, ws, headers, col_name, title):
		"""Return the value of `col_name` on the row whose title matches `title`."""
		col = headers.index(col_name) + 1
		title_col = headers.index("title") + 1
		for r in range(2, ws.max_row + 1):
			if ws.cell(row=r, column=title_col).value == title:
				return ws.cell(row=r, column=col).value
		self.fail(f"Row with title {title!r} not found")

	def test_country_normalization_columns_present(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Multiple Sclerosis"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			for col in ("countries_by_source", "regions_normalized", "trial_countries"):
				self.assertIn(col, headers, f'Column "{col}" missing from header row')
		finally:
			os.unlink(path)

	def test_trial_countries_rendered(self):
		"""TrialCountry rows render as 'Name [CODE] (status; date; src: …)'."""
		# Bypass save() so sync_trial_countries() does not overwrite these rows.
		TrialCountry.objects.create(
			trial=self.trial_ms,
			country="DE",
			status="recruiting",
			status_raw="Ongoing, recruiting",
			decision_date="2024-07-19",
			sources=["ctgov", "ctis"],
		)
		TrialCountry.objects.create(
			trial=self.trial_ms,
			country="FR",
			sources=["ctgov"],
		)
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Multiple Sclerosis"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			value = self._cell_for_title(
				ws,
				headers,
				"trial_countries",
				"A randomised trial of natalizumab in MS",
			)
			self.assertIn("Germany [DE]", value)
			self.assertIn("Ongoing, recruiting", value)
			self.assertIn("2024-07-19", value)
			self.assertIn("src: ctgov+ctis", value)
			self.assertIn("France [FR]", value)
		finally:
			os.unlink(path)

	def test_regions_and_countries_by_source_rendered(self):
		Trials.objects.filter(pk=self.trial_ms.pk).update(
			regions_normalized=["europe", "north_america"],
			countries_by_source={"ctgov": "Germany, United States", "ictrp": "France"},
		)
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Multiple Sclerosis"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			regions = self._cell_for_title(
				ws, headers, "regions_normalized",
				"A randomised trial of natalizumab in MS",
			)
			# list-of-scalars renders as a "; "-joined string, not a Python repr
			self.assertEqual(regions, "europe; north_america")

			by_source = self._cell_for_title(
				ws, headers, "countries_by_source",
				"A randomised trial of natalizumab in MS",
			)
			self.assertIn("ctgov", by_source)
			self.assertIn("United States", by_source)
			self.assertIn("ictrp", by_source)
		finally:
			os.unlink(path)

	def test_country_columns_have_glossary_entries(self):
		"""New columns must be documented (non-empty description) in the Glossary."""
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Glossary"]
			described = {}
			for r in range(2, ws.max_row + 1):
				described[ws.cell(row=r, column=1).value] = ws.cell(
					row=r, column=3
				).value
			for col in ("countries_by_source", "regions_normalized", "trial_countries"):
				self.assertIn(col, described)
				self.assertTrue(
					described[col], f'Glossary description for "{col}" is empty'
				)
		finally:
			os.unlink(path)

	# ------------------------------------------------------------------
	# Normalized sponsor columns
	# ------------------------------------------------------------------

	def test_sponsor_normalization_columns_present(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Multiple Sclerosis"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			for col in (
				"lead_sponsor_class",
				"sponsor_id",
				"sponsor_slug",
				"primary_sponsor_normalized",
				"sponsor_type_normalized",
				"sponsor_type_source",
			):
				self.assertIn(col, headers, f'Column "{col}" missing from header row')
		finally:
			os.unlink(path)

	def test_sponsor_id_and_slug_render_canonical_entity_identifiers(self):
		Trials.objects.filter(pk=self.trial_ms.pk).update(
			primary_sponsor="Novartis Pharma AG"
		)
		self.trial_ms.refresh_from_db()
		self.trial_ms.save()  # trigger sponsor resolution
		self.trial_ms.refresh_from_db()
		sponsor = self.trial_ms.primary_sponsor_normalized
		self.assertIsNotNone(sponsor)

		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Multiple Sclerosis"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			id_value = self._cell_for_title(
				ws, headers, "sponsor_id", "A randomised trial of natalizumab in MS"
			)
			slug_value = self._cell_for_title(
				ws, headers, "sponsor_slug", "A randomised trial of natalizumab in MS"
			)
			self.assertEqual(id_value, sponsor.pk)
			self.assertEqual(slug_value, sponsor.slug)
		finally:
			os.unlink(path)

	def test_primary_sponsor_normalized_renders_canonical_name(self):
		Trials.objects.filter(pk=self.trial_ms.pk).update(
			primary_sponsor="Novartis Pharma AG"
		)
		self.trial_ms.refresh_from_db()
		self.trial_ms.save()  # trigger sponsor resolution
		self.trial_ms.refresh_from_db()
		self.assertIsNotNone(self.trial_ms.primary_sponsor_normalized)

		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Multiple Sclerosis"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			value = self._cell_for_title(
				ws,
				headers,
				"primary_sponsor_normalized",
				"A randomised trial of natalizumab in MS",
			)
			self.assertEqual(value, self.trial_ms.primary_sponsor_normalized.name)
		finally:
			os.unlink(path)

	def test_sponsor_type_normalized_renders_derived_type(self):
		Trials.objects.filter(pk=self.trial_ms.pk).update(
			primary_sponsor="Acme Pharmaceuticals Inc.", lead_sponsor_class="INDUSTRY"
		)
		self.trial_ms.refresh_from_db()
		self.trial_ms.save()  # trigger sponsor resolution + type derivation
		self.trial_ms.refresh_from_db()

		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Multiple Sclerosis"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			value = self._cell_for_title(
				ws,
				headers,
				"sponsor_type_normalized",
				"A randomised trial of natalizumab in MS",
			)
			self.assertEqual(value, "industry")
		finally:
			os.unlink(path)

	def test_sponsor_type_source_shows_derivation_provenance(self):
		"""sponsor_type_source is the audit trail for sponsor_type_normalized: it must
		show which signal actually won (here, CTGov's lead_sponsor_class, since that
		outranks the name-keyword-rules tier a name like this would otherwise fall to)."""
		Trials.objects.filter(pk=self.trial_ms.pk).update(
			primary_sponsor="Acme Pharmaceuticals Inc.", lead_sponsor_class="INDUSTRY"
		)
		self.trial_ms.refresh_from_db()
		self.trial_ms.save()
		self.trial_ms.refresh_from_db()

		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Multiple Sclerosis"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			value = self._cell_for_title(
				ws,
				headers,
				"sponsor_type_source",
				"A randomised trial of natalizumab in MS",
			)
			self.assertEqual(value, "ctgov")
		finally:
			os.unlink(path)

	def test_unresolved_sponsor_columns_render_blank(self):
		"""A trial with no primary_sponsor must not error and must render blank, not
		crash on a None FK dereference."""
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Multiple Sclerosis"]
			headers = [
				ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
			]
			# openpyxl round-trips an empty string as None on some versions — either is
			# an acceptable "blank" rendering, matching the codebase's existing
			# `cell_val or ""` convention (see test_articles_column_shows_link).
			for col in (
				"sponsor_id",
				"sponsor_slug",
				"primary_sponsor_normalized",
				"sponsor_type_normalized",
				"sponsor_type_source",
			):
				value = self._cell_for_title(
					ws, headers, col, "A randomised trial of natalizumab in MS"
				)
				self.assertFalse(value, f'Expected "{col}" to render blank, got {value!r}')
		finally:
			os.unlink(path)

	def test_sponsor_columns_have_glossary_entries(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Glossary"]
			described = {}
			for r in range(2, ws.max_row + 1):
				described[ws.cell(row=r, column=1).value] = ws.cell(
					row=r, column=3
				).value
			for col in (
				"lead_sponsor_class",
				"sponsor_id",
				"sponsor_slug",
				"primary_sponsor_normalized",
				"sponsor_type_normalized",
				"sponsor_type_source",
			):
				self.assertIn(col, described)
				self.assertTrue(
					described[col], f'Glossary description for "{col}" is empty'
				)
		finally:
			os.unlink(path)

	# ------------------------------------------------------------------
	# Glossary sheet
	# ------------------------------------------------------------------

	def test_glossary_row_count_matches_data_columns(self):
		"""Every exported column must have exactly one Glossary row."""
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws_data = wb["Multiple Sclerosis"]
			ws_glossary = wb["Glossary"]
			data_col_count = ws_data.max_column
			# Glossary row 1 is the header; rows 2..N are one per column
			glossary_data_rows = ws_glossary.max_row - 1
			self.assertEqual(data_col_count, glossary_data_rows)
		finally:
			os.unlink(path)

	def test_glossary_has_field_and_label_columns(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Glossary"]
			self.assertEqual(ws.cell(row=1, column=1).value, "Field")
			self.assertEqual(ws.cell(row=1, column=2).value, "Label")
		finally:
			os.unlink(path)

	# ------------------------------------------------------------------
	# Registries sheet
	# ------------------------------------------------------------------

	def test_registries_sheet_has_three_registries(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Registries"]
			# Read all text in the sheet
			all_values = set()
			for row in ws.iter_rows(values_only=True):
				for v in row:
					if v:
						all_values.add(str(v))
			for reg in REGISTRY_NAMES:
				found = any(reg in v for v in all_values)
				self.assertTrue(
					found, f'Registry "{reg}" not found in Registries sheet'
				)
		finally:
			os.unlink(path)

	def test_registries_sheet_has_field_matrix(self):
		"""The field matrix header row must contain the registry names."""
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Registries"]
			all_values = []
			for row in ws.iter_rows(values_only=True):
				all_values.extend(v for v in row if v)
			for reg in REGISTRY_NAMES:
				self.assertIn(
					reg, all_values, f'Registry column "{reg}" not in field matrix'
				)
		finally:
			os.unlink(path)

	# ------------------------------------------------------------------
	# Categories sheet
	# ------------------------------------------------------------------

	def _category_row(self, ws, subject_name, category_name):
		"""Return {header: value} for the Categories sheet row matching (subject, category)."""
		headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
		subj_col = headers.index("Subject") + 1
		cat_col = headers.index("Category") + 1
		for r in range(5, ws.max_row + 1):
			if (
				ws.cell(row=r, column=subj_col).value == subject_name
				and ws.cell(row=r, column=cat_col).value == category_name
			):
				return {h: ws.cell(row=r, column=c + 1).value for c, h in enumerate(headers)}
		self.fail(f"Row for subject={subject_name!r} category={category_name!r} not found")

	def test_categories_sheet_present(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			self.assertIn("Categories", wb.sheetnames)
		finally:
			os.unlink(path)

	def test_categories_sheet_headers(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Categories"]
			headers = [
				ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)
			]
			self.assertEqual(headers, CATEGORY_COLUMNS)
		finally:
			os.unlink(path)

	def test_category_description_and_terms_rendered(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Categories"]
			row = self._category_row(ws, "Multiple Sclerosis", "Natalizumab")
			self.assertEqual(
				row["Description"], "Trials studying natalizumab (Tysabri) in MS."
			)
			self.assertEqual(row["Search terms"], "natalizumab; tysabri")
		finally:
			os.unlink(path)

	def test_category_row_per_subject_pair(self):
		path, wb = self._export(
			subjects=f"{self.subject_ms.pk},{self.subject_cancer.pk}"
		)
		try:
			ws = wb["Categories"]
			# Present under both subjects it is assigned to
			self._category_row(ws, "Multiple Sclerosis", "Shared category")
			self._category_row(ws, "Cancer", "Shared category")
		finally:
			os.unlink(path)

	def test_category_trial_count_scoped_to_subject(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Categories"]
			row = self._category_row(ws, "Multiple Sclerosis", "Natalizumab")
			self.assertEqual(row["Trials (this subject)"], 1)
		finally:
			os.unlink(path)

	def test_manual_category_shows_type_and_empty_terms(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Categories"]
			row = self._category_row(ws, "Multiple Sclerosis", "Manual watchlist")
			self.assertEqual(row["Category type"], "Manual")
			self.assertFalse(row["Search terms"])
		finally:
			os.unlink(path)

	def test_match_config_columns_populated(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Categories"]
			row = self._category_row(ws, "Multiple Sclerosis", "Natalizumab")
			self.assertTrue(row["Match scope"])
			self.assertTrue(row["Min score (trials)"])
			self.assertTrue(row["Field weights (trials)"])
		finally:
			os.unlink(path)

	def test_category_description_formula_injection_defused(self):
		"""A description starting with =, +, -, or @ must not become a live formula cell."""
		cat_formula = TeamCategory.objects.create(
			team=self.team,
			category_name="Formula-like description",
			category_description="=SUM(A1:A10)",
			category_terms=["+1(555)", "-benchmark", "@mention"],
		)
		cat_formula.subjects.add(self.subject_ms)
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Categories"]
			headers = [
				ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)
			]
			cat_col = headers.index("Category") + 1
			desc_col = headers.index("Description") + 1
			for r in range(5, ws.max_row + 1):
				if ws.cell(row=r, column=cat_col).value == "Formula-like description":
					cell = ws.cell(row=r, column=desc_col)
					self.assertNotEqual(cell.data_type, "f")
					self.assertEqual(cell.value, "'=SUM(A1:A10)")
					break
			else:
				self.fail("Row for Formula-like description not found")
		finally:
			os.unlink(path)

	def test_subjectless_category_never_exported(self):
		path, wb = self._export(
			subjects=f"{self.subject_ms.pk},{self.subject_cancer.pk}"
		)
		try:
			ws = wb["Categories"]
			names = [ws.cell(row=r, column=3).value for r in range(5, ws.max_row + 1)]
			self.assertNotIn("Orphan category", names)
		finally:
			os.unlink(path)

		path, wb = self._export(all_subjects=True)
		try:
			ws = wb["Categories"]
			names = [ws.cell(row=r, column=3).value for r in range(5, ws.max_row + 1)]
			self.assertNotIn("Orphan category", names)
		finally:
			os.unlink(path)

	def test_categories_sheet_explains_matching(self):
		path, wb = self._export(subjects=str(self.subject_ms.pk))
		try:
			ws = wb["Categories"]
			self.assertTrue(ws.cell(row=1, column=1).value)
			prose = ws.cell(row=2, column=1).value
			self.assertTrue(prose)
			self.assertIn("score", prose.lower())
		finally:
			os.unlink(path)

	def test_no_categories_writes_placeholder(self):
		no_cat_subject = Subject.objects.create(
			subject_name="No Categories Subject",
			subject_slug="no-categories-subject",
			team=self.team,
		)
		path, wb = self._export(subjects=str(no_cat_subject.pk))
		try:
			ws = wb["Categories"]
			self.assertEqual(ws.cell(row=5, column=1).value, "No categories found.")
		finally:
			os.unlink(path)

	def test_reserved_sheet_names(self):
		clash = Subject.objects.create(
			subject_name="Categories", subject_slug="categories-clash", team=self.team
		)
		path, wb = self._export(subjects=str(clash.pk))
		try:
			self.assertIn("Categories", wb.sheetnames)
			self.assertIn("Categories_2", wb.sheetnames)
			ws = wb["Categories"]
			headers = [
				ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)
			]
			self.assertEqual(headers, CATEGORY_COLUMNS)
		finally:
			os.unlink(path)

	# ------------------------------------------------------------------
	# Helper / utility unit tests
	# ------------------------------------------------------------------

	def test_sanitise_sheet_name_strips_illegal_chars(self):
		used = set()
		name = _sanitise_sheet_name("Trial/Results [2024]", used)
		self.assertNotIn("/", name)
		self.assertNotIn("[", name)
		self.assertNotIn("]", name)

	def test_sanitise_sheet_name_truncates_to_31(self):
		used = set()
		long_name = "A" * 50
		result = _sanitise_sheet_name(long_name, used)
		self.assertLessEqual(len(result), 31)

	def test_sanitise_sheet_name_deduplicates(self):
		used = set()
		name1 = _sanitise_sheet_name("Duplicate", used)
		name2 = _sanitise_sheet_name("Duplicate", used)
		self.assertNotEqual(name1, name2)

	def test_build_scalar_columns_excludes_generated_fields(self):
		scalar_cols = _build_scalar_columns()
		self.assertNotIn("utitle", scalar_cols)
		self.assertNotIn("usummary", scalar_cols)

	def test_build_scalar_columns_excludes_identifiers(self):
		scalar_cols = _build_scalar_columns()
		self.assertNotIn("identifiers", scalar_cols)

	def test_build_scalar_columns_includes_all_scalar_fields(self):
		scalar_cols = _build_scalar_columns()
		# Spot-check a broad cross-section of known fields
		for col in [
			"title",
			"summary",
			"phase",
			"condition",
			"ctg_detailed_description",
			"therapeutic_areas",
			"sponsor_type",
			"ethics_review_status",
		]:
			self.assertIn(col, scalar_cols, f'Expected column "{col}" in scalar list')


class AboutSheetTests(TestCase):
	"""Tests for the 'About this file' sheet and its --site resolution."""

	def setUp(self):
		self.org = Organization.objects.create(name="About Org", slug="about-org")
		self.site = Site.objects.create(
			domain="about-test.example", name="About Test Site"
		)
		self.custom_setting = CustomSetting.objects.create(
			site=self.site,
			title="About Test Project",
			admin_email="admin@about-test.example",
			website_url="https://about-test.example/",
			about_url="https://about-test.example/about/",
			contact_url="https://about-test.example/contact/",
			api_domain="api.about-test.example",
			github_url="https://github.com/example/about-test",
		)
		self.team = Team.objects.create(
			name="About Team", organization=self.org, slug="about-team", site=self.site
		)
		self.subject = Subject.objects.create(
			subject_name="About Subject",
			subject_slug="about-subject",
			team=self.team,
			description="Subject description text.",
		)
		self.trial = Trials.objects.create(
			title="A trial for the about sheet",
			link="https://clinicaltrials.gov/ct2/show/NCT00222222",
			identifiers={"nct": "NCT00222222"},
		)
		self.trial.subjects.add(self.subject)

	def _export(self, **kwargs):
		"""Run the command into a temp file; return (path, workbook)."""
		fd, path = tempfile.mkstemp(suffix=".xlsx")
		os.close(fd)
		call_command("export_trials_xlsx", output=path, **kwargs)
		wb = openpyxl.load_workbook(path)
		return path, wb

	def _about_rows(self, ws):
		"""Return {label: value} for the About sheet's label/value rows."""
		rows = {}
		for r in range(1, ws.max_row + 1):
			label = ws.cell(row=r, column=1).value
			value = ws.cell(row=r, column=2).value
			if label:
				rows[label] = value
		return rows

	# ------------------------------------------------------------------
	# Sheet placement
	# ------------------------------------------------------------------

	def test_about_sheet_is_first(self):
		path, wb = self._export(subjects=str(self.subject.pk))
		try:
			self.assertEqual(wb.sheetnames[0], "About this file")
		finally:
			os.unlink(path)

	# ------------------------------------------------------------------
	# --site flag
	# ------------------------------------------------------------------

	def test_site_flag_by_numeric_id(self):
		path, wb = self._export(subjects=str(self.subject.pk), site=str(self.site.pk))
		try:
			rows = self._about_rows(wb["About this file"])
			self.assertEqual(rows["Published by"], "About Test Project")
		finally:
			os.unlink(path)

	def test_site_flag_by_domain_case_insensitive(self):
		path, wb = self._export(
			subjects=str(self.subject.pk), site=self.site.domain.upper()
		)
		try:
			rows = self._about_rows(wb["About this file"])
			self.assertEqual(rows["Published by"], "About Test Project")
		finally:
			os.unlink(path)

	def test_unknown_site_flag_raises_naming_valid_sites(self):
		with self.assertRaises(CommandError) as ctx:
			call_command(
				"export_trials_xlsx",
				subjects=str(self.subject.pk),
				site="does-not-exist.example",
				output="/tmp/unused-about-sheet-test.xlsx",
			)
		self.assertIn(self.site.domain, str(ctx.exception))

	# ------------------------------------------------------------------
	# Default resolution
	# ------------------------------------------------------------------

	def test_team_site_wins_over_org_default(self):
		other_site = Site.objects.create(
			domain="org-default.example", name="Org Default Site"
		)
		OrganizationSite.objects.create(
			organization=self.org, site=other_site, is_default=True
		)
		# self.team.site is already set to self.site, which must win.
		path, wb = self._export(subjects=str(self.subject.pk))
		try:
			rows = self._about_rows(wb["About this file"])
			self.assertEqual(rows["Published by"], "About Test Project")
		finally:
			os.unlink(path)

	def test_org_default_used_when_team_site_is_none(self):
		team_no_site = Team.objects.create(
			name="No Site Team", organization=self.org, slug="no-site-team"
		)
		OrganizationSite.objects.create(
			organization=self.org, site=self.site, is_default=True
		)
		subject = Subject.objects.create(
			subject_name="Org Default Subject",
			subject_slug="org-default-subject",
			team=team_no_site,
		)
		trial = Trials.objects.create(
			title="Trial for org default site",
			link="https://clinicaltrials.gov/ct2/show/NCT00333333",
			identifiers={"nct": "NCT00333333"},
		)
		trial.subjects.add(subject)
		path, wb = self._export(subjects=str(subject.pk))
		try:
			rows = self._about_rows(wb["About this file"])
			self.assertEqual(rows["Published by"], "About Test Project")
		finally:
			os.unlink(path)

	def test_no_site_configured_writes_placeholder_and_does_not_raise(self):
		# Point SITE_ID at a row that doesn't exist so resolution has nothing left
		# to fall back to: no Team.site, no OrganizationSite default, and no current
		# Site. Uses override_settings rather than deleting the real Site(pk=1) row,
		# which would invalidate Django's process-wide site cache and leak into
		# unrelated tests' query-count assertions.
		org = Organization.objects.create(name="No Site Org", slug="no-site-org")
		team = Team.objects.create(
			name="No Site Team", organization=org, slug="no-site-team"
		)
		subject = Subject.objects.create(
			subject_name="No Site Subject", subject_slug="no-site-subject", team=team
		)
		trial = Trials.objects.create(
			title="Trial with no resolvable site",
			link="https://clinicaltrials.gov/ct2/show/NCT00444444",
			identifiers={"nct": "NCT00444444"},
		)
		trial.subjects.add(subject)
		with self.settings(SITE_ID=999999999):
			path, wb = self._export(subjects=str(subject.pk))
		try:
			ws = wb["About this file"]
			col_a_values = [
				ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)
			]
			self.assertTrue(
				any(
					v and "No site is configured" in str(v)
					for v in col_a_values
				)
			)
		finally:
			os.unlink(path)

	def test_multiple_sites_majority_wins_with_warning_and_note(self):
		minority_site = Site.objects.create(
			domain="minority-site.example", name="Minority Site"
		)
		minority_team = Team.objects.create(
			name="Minority Team",
			organization=self.org,
			slug="minority-team",
			site=minority_site,
		)
		minority_subject = Subject.objects.create(
			subject_name="Minority Subject",
			subject_slug="minority-subject",
			team=minority_team,
		)
		minority_trial = Trials.objects.create(
			title="Minority site trial",
			link="https://clinicaltrials.gov/ct2/show/NCT00555555",
			identifiers={"nct": "NCT00555555"},
		)
		minority_trial.subjects.add(minority_subject)

		out = StringIO()
		path, wb = self._export(
			subjects=f"{self.subject.pk},{minority_subject.pk}", stdout=out
		)
		try:
			self.assertIn("multiple sites", out.getvalue())
			rows = self._about_rows(wb["About this file"])
			# self.site backs one subject, minority_site backs one too — tie-break
			# on lowest pk decides which wins, so assert against whichever pk is lower.
			winner = self.site if self.site.pk < minority_site.pk else minority_site
			loser = minority_site if winner is self.site else self.site
			self.assertIn(loser.domain, out.getvalue())
			ws = wb["About this file"]
			col_a_values = [
				ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)
			]
			note_row = next(
				(v for v in col_a_values if v and "also contains subjects" in str(v)),
				None,
			)
			self.assertIsNotNone(note_row)
			self.assertIn(loser.domain, note_row)
		finally:
			os.unlink(path)

	def test_site_without_custom_setting_falls_back_to_site_name(self):
		bare_site = Site.objects.create(domain="bare-site.example", name="Bare Site")
		team = Team.objects.create(
			name="Bare Team", organization=self.org, slug="bare-team", site=bare_site
		)
		subject = Subject.objects.create(
			subject_name="Bare Subject", subject_slug="bare-subject", team=team
		)
		trial = Trials.objects.create(
			title="Trial for bare site",
			link="https://clinicaltrials.gov/ct2/show/NCT00666666",
			identifiers={"nct": "NCT00666666"},
		)
		trial.subjects.add(subject)
		path, wb = self._export(subjects=str(subject.pk))
		try:
			rows = self._about_rows(wb["About this file"])
			self.assertEqual(rows["Published by"], "Bare Site")
			self.assertEqual(rows["Website"], "https://bare-site.example")
		finally:
			os.unlink(path)

	# ------------------------------------------------------------------
	# Contact email fallback
	# ------------------------------------------------------------------

	def test_contact_email_blank_falls_back_to_admin_email(self):
		path, wb = self._export(subjects=str(self.subject.pk))
		try:
			rows = self._about_rows(wb["About this file"])
			self.assertEqual(rows["Contact email"], "admin@about-test.example")
		finally:
			os.unlink(path)

	def test_contact_email_set_wins_over_admin_email(self):
		self.custom_setting.contact_email = "public@about-test.example"
		self.custom_setting.save()
		path, wb = self._export(subjects=str(self.subject.pk))
		try:
			rows = self._about_rows(wb["About this file"])
			self.assertEqual(rows["Contact email"], "public@about-test.example")
		finally:
			os.unlink(path)

	# ------------------------------------------------------------------
	# Citation
	# ------------------------------------------------------------------

	def test_blank_citation_is_generated_with_title_and_year(self):
		path, wb = self._export(subjects=str(self.subject.pk))
		try:
			rows = self._about_rows(wb["About this file"])
			citation = rows["How to cite"]
			self.assertIn("About Test Project", citation)
			import datetime as dt

			self.assertIn(str(dt.date.today().year), citation)
		finally:
			os.unlink(path)

	def test_set_citation_is_used_verbatim(self):
		self.custom_setting.citation = "Custom citation string."
		self.custom_setting.save()
		path, wb = self._export(subjects=str(self.subject.pk))
		try:
			rows = self._about_rows(wb["About this file"])
			self.assertEqual(rows["How to cite"], "Custom citation string.")
		finally:
			os.unlink(path)

	# ------------------------------------------------------------------
	# Blank fields produce no row
	# ------------------------------------------------------------------

	def test_blank_field_produces_no_row(self):
		# privacy_policy_url is blank by default in setUp's CustomSetting.
		path, wb = self._export(subjects=str(self.subject.pk))
		try:
			rows = self._about_rows(wb["About this file"])
			self.assertNotIn("Privacy policy", rows)
		finally:
			os.unlink(path)

	# ------------------------------------------------------------------
	# Formula injection
	# ------------------------------------------------------------------

	def test_description_formula_injection_defused(self):
		self.custom_setting.description = "=SUM(A1:A10)"
		self.custom_setting.save()
		path, wb = self._export(subjects=str(self.subject.pk))
		try:
			ws = wb["About this file"]
			rows = {}
			for r in range(1, ws.max_row + 1):
				if ws.cell(row=r, column=1).value == "About this project":
					cell = ws.cell(row=r, column=2)
					self.assertNotEqual(cell.data_type, "f")
					self.assertEqual(cell.value, "'=SUM(A1:A10)")
					break
			else:
				self.fail("About this project row not found")
		finally:
			os.unlink(path)

	# ------------------------------------------------------------------
	# Reserved sheet name
	# ------------------------------------------------------------------

	def test_subject_named_about_this_file_gets_suffixed_sheet(self):
		clash = Subject.objects.create(
			subject_name="About this file",
			subject_slug="about-this-file-clash",
			team=self.team,
		)
		trial = Trials.objects.create(
			title="Trial for the clashing subject",
			link="https://clinicaltrials.gov/ct2/show/NCT00777777",
			identifiers={"nct": "NCT00777777"},
		)
		trial.subjects.add(clash)
		path, wb = self._export(subjects=str(clash.pk))
		try:
			self.assertEqual(wb.sheetnames[0], "About this file")
			self.assertIn("About this file_2", wb.sheetnames)
		finally:
			os.unlink(path)

	# ------------------------------------------------------------------
	# Section 3 — What's in this workbook
	# ------------------------------------------------------------------

	def test_whats_in_this_workbook_lists_every_subject_sheet_with_count(self):
		path, wb = self._export(subjects=str(self.subject.pk))
		try:
			rows = self._about_rows(wb["About this file"])
			self.assertIn("About Subject", rows)
			self.assertIn("1 clinical trial(s)", rows["About Subject"])
			self.assertIn("Subject description text.", rows["About Subject"])
			self.assertIn("Categories", rows)
			self.assertIn("Glossary", rows)
			self.assertIn("Registries", rows)
		finally:
			os.unlink(path)
